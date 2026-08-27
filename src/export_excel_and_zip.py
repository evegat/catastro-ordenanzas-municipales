"""
Generador de entregables descargables:
1. Excel profesional (.xlsx) con 1.632 ordenanzas, hipervínculos, resumen comunal y temático.
2. Paquete ZIP consolidado para descarga directa desde el dashboard web.
"""

import os
import csv
import json
import zipfile
import sqlite3
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = Path("D:/Datasets/P090 - BCN Ordenanzas municipales/catastro_ordenanzas.db")
MAESTRO_CSV = Path("D:/Proyectos/P090 - Catastro Ordenanzas Municipales BCN/data/maestro_comunas_chile.csv")
DESCARGAS_DIR = Path("D:/Proyectos/P090 - Catastro Ordenanzas Municipales BCN/dashboard/descargas")
EXCEL_OUTPUT = DESCARGAS_DIR / "catastro_ordenanzas_nacional_2026.xlsx"
ZIP_OUTPUT = DESCARGAS_DIR / "consolidado_ordenanzas_chile_2026.zip"

def build_excel_export():
    DESCARGAS_DIR.mkdir(parents=True, exist_ok=True)

    # Load Communes Maestro for Region Mapping
    commune_reg_map = {}
    if MAESTRO_CSV.exists():
        with open(MAESTRO_CSV, mode="r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                commune_reg_map[r["comuna_nombre"].lower().strip()] = {
                    "region_id": r["region_id"],
                    "region_nombre": r["region_nombre"]
                }

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT id, fuente, municipalidad_slug, comuna_nombre, fecha_publicacion, numero, titulo, materia, pdf_url, estado
        FROM ordenanzas
        ORDER BY fecha_publicacion DESC
    """)
    rows = cursor.fetchall()
    conn.close()

    print(f"Exportando {len(rows)} ordenanzas a Excel...")

    wb = openpyxl.Workbook()

    # Define Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="475569")
    bold_font = Font(name="Calibri", size=10, bold=True)
    regular_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="0284C7", underline="single")
    cplt_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
    bcn_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # -------------------------------------------------------------
    # SHEET 1: CATASTRO COMPLETO DE ORDENANZAS
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Catastro de Ordenanzas"

    # Title block
    ws1["A1"] = "P090 — Catastro Nacional de Ordenanzas Municipales de Chile"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Elaborado por Eduardo Vega • Consolidado BCN LeyChile & Transparencia Activa CPLT • Actualizado al {datetime.now().strftime('%d/%m/%Y')}"
    ws1["A2"].font = subtitle_font

    headers_1 = [
        "ID", "Región N°", "Región", "Comuna", "Fuente", "N° Decreto / Ordenanza", 
        "Fecha", "Año", "Materia / Área Temática", "Título Oficial de la Ordenanza", "Enlace Documento Oficial", "Estado"
    ]

    for col_num, h_text in enumerate(headers_1, 1):
        cell = ws1.cell(row=4, column=col_num, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    topic_counts = {}
    commune_counts = {}

    for row_idx, r in enumerate(rows, 5):
        db_id, fuente, slug, comuna, fecha, num, titulo, materia, pdf_url, estado = r
        
        reg_info = commune_reg_map.get(comuna.lower().strip(), {"region_id": "—", "region_nombre": "—"})
        anio = fecha[:4] if fecha and len(fecha) >= 4 else "—"

        topic_counts[materia] = topic_counts.get(materia, 0) + 1
        commune_counts[comuna] = commune_counts.get(comuna, 0) + 1

        fuente_desc = "Transparencia CPLT (2022-26)" if fuente == "CPLT" else "BCN LeyChile"

        row_data = [
            db_id,
            reg_info["region_id"],
            reg_info["region_nombre"],
            comuna,
            fuente_desc,
            num or "S/N",
            fecha or "—",
            anio,
            materia or "General",
            titulo,
            pdf_url or "—",
            estado or "REGISTRADO"
        ]

        for col_num, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_num, value=val)
            cell.font = regular_font
            cell.border = thin_border

            if col_num in [1, 2, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = cplt_fill if fuente == "CPLT" else bcn_fill
            elif col_num == 11 and str(val).startswith("http"):
                cell.font = link_font
                cell.hyperlink = val

    # Auto-fit columns
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 65)

    ws1.auto_filter.ref = f"A4:L{len(rows)+4}"
    ws1.freeze_panes = "A5"

    # -------------------------------------------------------------
    # SHEET 2: RESUMEN POR ÁREAS TEMÁTICAS
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Distribución Temática")
    ws2["A1"] = "Distribución Semántica de Ordenanzas Municipales"
    ws2["A1"].font = title_font

    headers_2 = ["Área Temática / Materia", "Cantidad de Ordenanzas", "% del Total"]
    for col_num, h_text in enumerate(headers_2, 1):
        cell = ws2.cell(row=3, column=col_num, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_norms = len(rows)
    for r_idx, (top_name, count) in enumerate(sorted(topic_counts.items(), key=lambda x: x[1], reverse=True), 4):
        pct = (count / total_norms) * 100
        ws2.cell(row=r_idx, column=1, value=top_name).font = bold_font
        ws2.cell(row=r_idx, column=2, value=count).alignment = Alignment(horizontal="center")
        ws2.cell(row=r_idx, column=3, value=f"{pct:.1f}%").alignment = Alignment(horizontal="center")
        for c in range(1, 4):
            ws2.cell(row=r_idx, column=c).border = thin_border

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 18

    # -------------------------------------------------------------
    # SHEET 3: RESUMEN DE COBERTURA COMUNAL
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Cobertura por Comuna")
    ws3["A1"] = "Catastro Comunal — 346 Comunas de Chile"
    ws3["A1"].font = title_font

    headers_3 = ["Región", "Comuna", "Total Ordenanzas", "Estado de Cobertura"]
    for col_num, h_text in enumerate(headers_3, 1):
        cell = ws3.cell(row=3, column=col_num, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if MAESTRO_CSV.exists():
        with open(MAESTRO_CSV, mode="r", encoding="utf-8") as f:
            all_comunas = list(csv.DictReader(f))

        for r_idx, c in enumerate(all_comunas, 4):
            c_name = c["comuna_nombre"]
            cnt = commune_counts.get(c_name, 0)
            status = "Cobertura Completa" if cnt > 0 else "Fase 2 CPLT"
            
            ws3.cell(row=r_idx, column=1, value=f"{c['region_id']} - {c['region_nombre']}")
            ws3.cell(row=r_idx, column=2, value=c_name).font = bold_font
            ws3.cell(row=r_idx, column=3, value=cnt).alignment = Alignment(horizontal="center")
            ws3.cell(row=r_idx, column=4, value=status).alignment = Alignment(horizontal="center")
            for col_i in range(1, 5):
                ws3.cell(row=r_idx, column=col_i).border = thin_border

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 25

    wb.save(EXCEL_OUTPUT)
    print(f"Excel generado exitosamente: {EXCEL_OUTPUT}")

def build_zip_package():
    print(f"Generando paquete ZIP consolidado: {ZIP_OUTPUT}...")
    
    readme_content = f"""P090 — Catastro Nacional de Ordenanzas Municipales de Chile
===========================================================
Elaborado por: Eduardo Vega
Fecha de consolidación: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Total Ordenanzas: 1.632 normas estructuradas (BCN LeyChile & Transparencia Activa CPLT)

Contenido del paquete:
1. catastro_ordenanzas_nacional_2026.xlsx: Planilla Excel oficial con 1.632 ordenanzas, clasificación en 9 ejes temáticos, hipervínculos a LeyChile y resoluciones CPLT.
2. catastro_ordenanzas.db: Base de datos relacional SQLite completa e indexada.
3. maestro_comunas_chile.csv: Directorio de las 346 comunas de Chile con URLs de portales de transparencia.
4. ordenanzas_municipales_bcn_catalogo.csv: Catálogo histórico BCN con URIs LOD.

Acceso a la plataforma web en vivo:
https://villas-bradley-supreme-bubble.trycloudflare.com
"""

    with zipfile.ZipFile(ZIP_OUTPUT, "w", zipfile.ZIP_DEFLATED) as zf:
        if EXCEL_OUTPUT.exists():
            zf.write(EXCEL_OUTPUT, arcname="catastro_ordenanzas_nacional_2026.xlsx")
        if DB_PATH.exists():
            zf.write(DB_PATH, arcname="catastro_ordenanzas.db")
        if MAESTRO_CSV.exists():
            zf.write(MAESTRO_CSV, arcname="maestro_comunas_chile.csv")
        
        bcn_cat = Path("D:/Datasets/P090 - BCN Ordenanzas municipales/ordenanzas_municipales_bcn_catalogo.csv")
        if bcn_cat.exists():
            zf.write(bcn_cat, arcname="ordenanzas_municipales_bcn_catalogo.csv")

        zf.writestr("LEAME_ENTREGABLE.txt", readme_content)

    print(f"Paquete ZIP generado exitosamente ({os.path.getsize(ZIP_OUTPUT) / (1024*1024):.2f} MB): {ZIP_OUTPUT}")

if __name__ == "__main__":
    build_excel_export()
    build_zip_package()
